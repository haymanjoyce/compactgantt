import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Any, Type, Dict, List
from datetime import datetime
import logging
from models.project import ProjectData
from models.frame import FrameConfig
from models.task import Task
from models.link import Link
from models.pipe import Pipe
from models.curtain import Curtain
from models.swimlane import Swimlane
from models.text_box import TextBox
from utils.conversion import internal_to_display_date, display_to_internal_date


class ExcelRepository:
    """Repository for Excel (XLSX) import/export functionality."""
    
    def save(self, file_path: str, project_data: ProjectData) -> None:
        """Export project data to Excel file with worksheets for each tab."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create worksheets for each tab
        self._create_layout_sheet(wb, project_data.frame_config)
        self._create_titles_sheet(wb, project_data.frame_config)
        self._create_timeline_sheet(wb, project_data.frame_config)
        self._create_typography_sheet(wb, project_data.chart_config)
        # Grid sheet deprecated - horizontal gridlines now in Layout sheet as "Row Dividers"
        self._create_tasks_sheet(wb, project_data.tasks)
        self._create_links_sheet(wb, project_data.links)
        self._create_swimlanes_sheet(wb, project_data.swimlanes)
        self._create_pipes_sheet(wb, project_data.pipes)
        self._create_curtains_sheet(wb, project_data.curtains)
        self._create_text_boxes_sheet(wb, project_data.text_boxes)
        
        wb.save(file_path)
    
    def load(self, file_path: str, project_data_cls: Type) -> ProjectData:
        """Import project data from Excel file."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Create new project instance
        project = project_data_cls()
        
        # Load data from each worksheet
        frame_config_data = {}
        
        # Load Layout sheet
        if "Layout" in wb.sheetnames:
            layout_data = self._read_key_value_sheet(wb["Layout"])
            frame_config_data.update(layout_data)
        
        # Load Titles sheet
        if "Titles" in wb.sheetnames:
            titles_data = self._read_key_value_sheet(wb["Titles"])
            frame_config_data.update(titles_data)
        
        # Load Timeline sheet (new structure)
        if "Timeline" in wb.sheetnames:
            timeline_data = self._read_key_value_sheet(wb["Timeline"])
            frame_config_data.update(timeline_data)
        
        # Backward compatibility: Load old Scales sheet if Timeline doesn't exist
        if "Timeline" not in wb.sheetnames and "Scales" in wb.sheetnames:
            scales_data = self._read_key_value_sheet(wb["Scales"])
            frame_config_data.update(scales_data)
        
        # Load Grid sheet
        if "Grid" in wb.sheetnames:
            grid_data = self._read_key_value_sheet(wb["Grid"])
            frame_config_data.update(grid_data)
        
        # Convert margins list to tuple if present
        if "margins" in frame_config_data and isinstance(frame_config_data["margins"], list):
            frame_config_data["margins"] = tuple(frame_config_data["margins"])
        
        # Create FrameConfig
        project.frame_config = FrameConfig(**frame_config_data)
        
        # Load Tasks sheet
        if "Tasks" in wb.sheetnames:
            project.tasks = self._read_tasks_sheet(wb["Tasks"])
        
        # Load Links sheet - using header-based mapping instead of positional arrays
        if "Links" in wb.sheetnames:
            project.links = self._read_links_sheet(wb["Links"])
        
        if "Swimlanes" in wb.sheetnames:
            project.swimlanes = self._read_swimlanes_sheet(wb["Swimlanes"])
        
        if "Pipes" in wb.sheetnames:
            project.pipes = self._read_pipes_sheet(wb["Pipes"])
        
        if "Curtains" in wb.sheetnames:
            project.curtains = self._read_curtains_sheet(wb["Curtains"])
        
        if "Text Boxes" in wb.sheetnames:
            project.text_boxes = self._read_text_boxes_sheet(wb["Text Boxes"])
        
        # Load Typography sheet
        if "Typography" in wb.sheetnames:
            typography_data = self._read_typography_sheet(wb["Typography"])
            # Update project's chart_config with loaded values
            for key, value in typography_data.items():
                if hasattr(project.chart_config, key):
                    setattr(project.chart_config, key, value)
        
        return project
    
    def _create_layout_sheet(self, wb: Workbook, frame_config: FrameConfig) -> None:
        """Create Layout worksheet with chart dimensions, margins, and rows."""
        ws = wb.create_sheet("Layout")
        
        # Header
        ws.append(["Field", "Value"])
        self._format_header_row(ws, 1)
        
        # Layout fields
        ws.append(["Outer Width", frame_config.outer_width])
        ws.append(["Outer Height", frame_config.outer_height])
        ws.append(["Number of Rows", frame_config.num_rows])
        ws.append(["Margin Top", frame_config.margins[0]])
        ws.append(["Margin Right", frame_config.margins[1]])
        ws.append(["Margin Bottom", frame_config.margins[2]])
        ws.append(["Margin Left", frame_config.margins[3]])
        ws.append(["Row Numbers", "Yes" if getattr(frame_config, 'show_row_numbers', False) else "No"])
        ws.append(["Row Dividers", "Yes" if frame_config.horizontal_gridlines else "No"])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_titles_sheet(self, wb: Workbook, frame_config: FrameConfig) -> None:
        """Create Titles worksheet with header and footer settings."""
        ws = wb.create_sheet("Titles")
        
        # Header
        ws.append(["Field", "Value"])
        self._format_header_row(ws, 1)
        
        # Title fields
        ws.append(["Header Height", frame_config.header_height])
        ws.append(["Header Text", frame_config.header_text])
        ws.append(["Footer Height", frame_config.footer_height])
        ws.append(["Footer Text", frame_config.footer_text])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_timeline_sheet(self, wb: Workbook, frame_config: FrameConfig) -> None:
        """Create Timeline worksheet with timeframe, scales, and vertical gridlines."""
        ws = wb.create_sheet("Timeline")
        
        # Header
        ws.append(["Field", "Value"])
        self._format_header_row(ws, 1)
        
        # Timeframe fields
        ws.append(["Chart Start Date", internal_to_display_date(frame_config.chart_start_date)])
        ws.append(["Chart End Date", internal_to_display_date(frame_config.chart_end_date)])
        
        # Scale fields
        ws.append(["Show Years", "Yes" if frame_config.show_years else "No"])
        ws.append(["Show Months", "Yes" if frame_config.show_months else "No"])
        ws.append(["Show Weeks", "Yes" if frame_config.show_weeks else "No"])
        ws.append(["Show Days", "Yes" if frame_config.show_days else "No"])
        
        # Vertical Gridline fields
        ws.append(["Vertical Gridline Years", "Yes" if frame_config.vertical_gridline_years else "No"])
        ws.append(["Vertical Gridline Months", "Yes" if frame_config.vertical_gridline_months else "No"])
        ws.append(["Vertical Gridline Weeks", "Yes" if frame_config.vertical_gridline_weeks else "No"])
        ws.append(["Vertical Gridline Days", "Yes" if frame_config.vertical_gridline_days else "No"])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_typography_sheet(self, wb: Workbook, chart_config) -> None:
        """Create Typography worksheet with font and alignment settings."""
        from config.chart_config import ChartConfig
        ws = wb.create_sheet("Typography")
        
        # Header
        ws.append(["Field", "Value"])
        self._format_header_row(ws, 1)
        
        # Font Family
        ws.append(["Font Family", chart_config.font_family])
        
        # Font Sizes
        ws.append(["Task Font Size", chart_config.task_font_size])
        ws.append(["Scale Font Size", chart_config.scale_font_size])
        ws.append(["Header & Footer Font Size", chart_config.header_footer_font_size])
        ws.append(["Row Number Font Size", chart_config.row_number_font_size])
        ws.append(["Text Box Font Size", chart_config.text_box_font_size])
        ws.append(["Swimlane Font Size", chart_config.swimlane_font_size])
        
        # Vertical Alignment Factors
        ws.append(["Scale Vertical Alignment Factor", chart_config.scale_vertical_alignment_factor])
        ws.append(["Task Vertical Alignment Factor", chart_config.task_vertical_alignment_factor])
        ws.append(["Row Number Vertical Alignment Factor", chart_config.row_number_vertical_alignment_factor])
        ws.append(["Header & Footer Vertical Alignment Factor", chart_config.header_footer_vertical_alignment_factor])
        ws.append(["Swimlane Top Vertical Alignment Factor", chart_config.swimlane_top_vertical_alignment_factor])
        ws.append(["Swimlane Bottom Vertical Alignment Factor", chart_config.swimlane_bottom_vertical_alignment_factor])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def _create_grid_sheet(self, wb: Workbook, frame_config: FrameConfig) -> None:
        """DEPRECATED: Grid sheet is no longer created.
        
        Horizontal gridlines are now saved in Layout sheet as "Row Dividers".
        Vertical gridlines and scales are in Timeline sheet.
        Show Row Numbers is in Layout sheet.
        
        This method is kept for reference but is not called.
        """
        # This method is deprecated and should not be called
        pass
    
    def _create_tasks_sheet(self, wb: Workbook, tasks: List[Task]) -> None:
        """Create Tasks worksheet with task data as a table.
        
        Only saves visible/editable fields that match the UI table columns.
        Redundant fields (Is Milestone, Label Alignment, Label Horizontal Offset, Label Text Colour)
        are excluded as they are either auto-calculated or have default values.
        """
        ws = wb.create_sheet("Tasks")
        
        # Headers - only include fields that are visible/editable in UI
        headers = ["ID", "Row", "Name", "Start Date", "Finish Date", "Label Content", "Label Placement", "Label Offset", "Fill Color"]
        ws.append(headers)
        self._format_header_row(ws, 1)
        
        # Task rows - only save visible/editable fields
        for task in tasks:
            # Use label_content if available, otherwise fall back to label_hide for backward compatibility
            label_content = task.label_content if hasattr(task, 'label_content') and task.label_content else ("None" if task.label_hide == "No" else "Name only")
            row = [
                task.task_id,
                task.row_number,
                task.task_name,
                internal_to_display_date(task.start_date),
                internal_to_display_date(task.finish_date),
                label_content,
                task.label_placement,
                int(task.label_horizontal_offset) if task.label_horizontal_offset else 0,
                task.fill_color
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if header == "Name":
                ws.column_dimensions[col_letter].width = 15
            elif header in ["Start Date", "Finish Date"]:
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = 10
    
    def _create_links_sheet(self, wb: Workbook, links: List[Link]) -> None:
        """Create Links worksheet."""
        ws = wb.create_sheet("Links")
        
        # Define headers explicitly (excluding Valid field as it's calculated)
        headers = ["ID", "From Task ID", "From Task Name", "To Task ID", "To Task Name", "Line Color", "Line Style", "Link Routing"]
        ws.append(headers)
        self._format_header_row(ws, 1)
        
        # Write link data (Valid field is excluded as it's calculated)
        for link in links:
            ws.append([
                str(link.link_id),
                str(link.from_task_id),
                link.from_task_name or "",
                str(link.to_task_id),
                link.to_task_name or "",
                link.line_color,
                link.line_style,
                link.link_routing
            ])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def _create_swimlanes_sheet(self, wb: Workbook, swimlanes: List[Swimlane]) -> None:
        """Create Swimlanes worksheet."""
        ws = wb.create_sheet("Swimlanes")
        ws.append(["ID", "Row Count", "Title", "Label Position"])
        self._format_header_row(ws, 1)
        
        for swimlane in swimlanes:
            # Use title if available, fall back to name for backward compatibility
            title = swimlane.title if hasattr(swimlane, 'title') else (swimlane.name if hasattr(swimlane, 'name') else "")
            label_position = swimlane.label_position if hasattr(swimlane, 'label_position') else "Bottom Right"
            ws.append([
                swimlane.swimlane_id,
                swimlane.row_count,
                title if title else "",
                label_position
            ])
    
    def _create_pipes_sheet(self, wb: Workbook, pipes: List[Pipe]) -> None:
        """Create Pipes worksheet."""
        ws = wb.create_sheet("Pipes")
        ws.append(["ID", "Date", "Color", "Name"])
        self._format_header_row(ws, 1)
        
        for pipe in pipes:
            date_display = internal_to_display_date(pipe.date) if pipe.date else ""
            ws.append([
                pipe.pipe_id,
                date_display,
                pipe.color if pipe.color else "red",
                pipe.name if pipe.name else ""
            ])
    
    def _create_curtains_sheet(self, wb: Workbook, curtains: List[Curtain]) -> None:
        """Create Curtains worksheet."""
        ws = wb.create_sheet("Curtains")
        ws.append(["ID", "Start Date", "End Date", "Color", "Name"])
        self._format_header_row(ws, 1)
        
        for curtain in curtains:
            start_date_display = internal_to_display_date(curtain.start_date) if curtain.start_date else ""
            end_date_display = internal_to_display_date(curtain.end_date) if curtain.end_date else ""
            ws.append([
                curtain.curtain_id,
                start_date_display,
                end_date_display,
                curtain.color if curtain.color else "red",
                curtain.name if curtain.name else ""
            ])
    
    def _read_pipes_sheet(self, ws) -> List[Pipe]:
        """Read Pipes worksheet and return list of Pipe objects."""
        pipes = []
        headers = {}
        
        # Read header row
        header_row = ws[1]
        for idx, cell in enumerate(header_row):
            headers[cell.value] = idx
        
        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=False):
            if not any(cell.value for cell in row):
                continue
            
            pipe_data = {}
            for header, col_idx in headers.items():
                cell = row[col_idx]
                value = cell.value if cell else None
                
                if header == "ID":
                    pipe_data["pipe_id"] = int(value) if value is not None else 0
                elif header == "Date":
                    if value:
                        if isinstance(value, datetime):
                            pipe_data["date"] = value.strftime("%Y-%m-%d")
                        else:
                            pipe_data["date"] = display_to_internal_date(str(value))
                    else:
                        pipe_data["date"] = ""
                elif header == "Colour" or header == "Color":
                    pipe_data["color"] = str(value) if value else "red"
                elif header == "Name":
                    pipe_data["name"] = str(value) if value else ""
            
            if pipe_data.get("pipe_id") and pipe_data.get("date"):
                try:
                    pipes.append(Pipe.from_dict(pipe_data))
                except (ValueError, KeyError) as e:
                    logging.warning(f"Skipping invalid pipe row: {e}")
        
        return pipes
    
    def _read_curtains_sheet(self, ws) -> List[Curtain]:
        """Read Curtains worksheet and return list of Curtain objects."""
        curtains = []
        headers = {}
        
        # Read header row
        header_row = ws[1]
        for idx, cell in enumerate(header_row):
            headers[cell.value] = idx
        
        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=False):
            if not any(cell.value for cell in row):
                continue
            
            curtain_data = {}
            for header, col_idx in headers.items():
                cell = row[col_idx]
                value = cell.value if cell else None
                
                if header == "ID":
                    curtain_data["curtain_id"] = int(value) if value is not None else 0
                elif header == "Start Date":
                    if value:
                        if isinstance(value, datetime):
                            curtain_data["start_date"] = value.strftime("%Y-%m-%d")
                        else:
                            curtain_data["start_date"] = display_to_internal_date(str(value))
                    else:
                        curtain_data["start_date"] = ""
                elif header == "End Date":
                    if value:
                        if isinstance(value, datetime):
                            curtain_data["end_date"] = value.strftime("%Y-%m-%d")
                        else:
                            curtain_data["end_date"] = display_to_internal_date(str(value))
                    else:
                        curtain_data["end_date"] = ""
                elif header == "Colour" or header == "Color":
                    curtain_data["color"] = str(value) if value else "red"
                elif header == "Name":
                    curtain_data["name"] = str(value) if value else ""
            
            if curtain_data.get("curtain_id") and curtain_data.get("start_date") and curtain_data.get("end_date"):
                try:
                    curtains.append(Curtain.from_dict(curtain_data))
                except (ValueError, KeyError) as e:
                    logging.warning(f"Skipping invalid curtain row: {e}")
        
        return curtains
    
    def _read_swimlanes_sheet(self, ws) -> List[Swimlane]:
        """Read Swimlanes worksheet and return list of Swimlane objects."""
        swimlanes = []
        headers = {}
        
        # Read header row
        header_row = ws[1]
        for idx, cell in enumerate(header_row):
            if cell.value:
                headers[cell.value] = idx
        
        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=False):
            if not any(cell.value for cell in row):
                continue
            
            swimlane_data = {}
            for idx, cell in enumerate(row):
                # Get header by column index
                header = None
                header_row_values = [c.value for c in ws[1]]
                if idx < len(header_row_values):
                    header = header_row_values[idx]
                
                value = cell.value
                if header == "ID":
                    swimlane_data["swimlane_id"] = int(value) if value else 0
                elif header == "Row Count":
                    swimlane_data["row_count"] = int(value) if value else 0
                elif header == "Name" or header == "Title":
                    # Backward compatibility: support both 'Name' and 'Title'
                    swimlane_data["title"] = str(value) if value else ""
                elif header == "Label Position":
                    swimlane_data["label_position"] = str(value) if value else "Bottom Right"
                # Backward compatibility: support old First Row/Last Row format
                elif header == "First Row":
                    swimlane_data["first_row"] = int(value) if value else 0
                elif header == "Last Row":
                    swimlane_data["last_row"] = int(value) if value else 0
            
            if swimlane_data.get("swimlane_id"):
                try:
                    # Swimlane.from_dict() handles backward compatibility for first_row/last_row
                    swimlanes.append(Swimlane.from_dict(swimlane_data))
                except (ValueError, KeyError) as e:
                    logging.warning(f"Skipping invalid swimlane row: {e}")
        
        return swimlanes
    
    def _create_text_boxes_sheet(self, wb: Workbook, text_boxes: List[TextBox]) -> None:
        """Create Text Boxes worksheet."""
        ws = wb.create_sheet("Text Boxes")
        ws.append(["ID", "X", "Y", "Width", "Height", "Text Align", "Vertical Align", "Text"])
        self._format_header_row(ws, 1)
        
        for textbox in text_boxes:
            ws.append([
                textbox.textbox_id,
                textbox.x,
                textbox.y,
                textbox.width,
                textbox.height,
                textbox.text_align,
                textbox.vertical_align,
                textbox.text
            ])
    
    def _format_header_row(self, ws, row_num: int) -> None:
        """Format header row with bold font and background color."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _read_key_value_sheet(self, ws) -> Dict[str, Any]:
        """Read a key-value pair worksheet (Layout, Titles, Timeline, Grid)."""
        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row[0] and row[1] is not None:
                key = str(row[0]).strip()
                value = row[1]
                
                # Convert value based on key
                if key in ["Outer Width", "Outer Height", "Number of Rows", "Header Height", "Footer Height",
                           "Margin Top", "Margin Right", "Margin Bottom", "Margin Left"]:
                    try:
                        value = int(value) if value is not None else 0
                    except (ValueError, TypeError):
                        value = 0
                elif key in ["Show Years", "Show Months", "Show Weeks", "Show Days",
                             "Horizontal Gridlines", "Show Row Gridlines", "Row Dividers",
                             "Vertical Gridline Years", "Vertical Gridline Months",
                             "Vertical Gridline Weeks", "Vertical Gridline Days", 
                             "Row Numbers", "Show Row Numbers"]:
                    value = str(value).strip().lower() in ["yes", "true", "1", 1, True]
                elif key in ["Chart Start Date", "Chart End Date"]:
                    # Convert from display format (dd/mm/yyyy) to internal format (yyyy-mm-dd)
                    # Handle both string dates and Excel date serial numbers
                    try:
                        if value:
                            if isinstance(value, datetime):
                                # Excel date serial number converted to datetime
                                value = value.strftime("%Y-%m-%d")
                            else:
                                # String date - convert from display format
                                value = display_to_internal_date(str(value))
                    except (ValueError, AttributeError):
                        pass  # Keep original value if conversion fails
                elif key in ["Header Text", "Footer Text"]:
                    value = str(value) if value is not None else ""
                
                # Map to frame_config field names
                field_map = {
                    "Outer Width": "outer_width",
                    "Outer Height": "outer_height",
                    "Number of Rows": "num_rows",
                    "Margin Top": "margins",
                    "Margin Right": "margins",
                    "Margin Bottom": "margins",
                    "Margin Left": "margins",
                    "Header Height": "header_height",
                    "Header Text": "header_text",
                    "Footer Height": "footer_height",
                    "Footer Text": "footer_text",
                    "Show Years": "show_years",
                    "Show Months": "show_months",
                    "Show Weeks": "show_weeks",
                    "Show Days": "show_days",
                    "Horizontal Gridlines": "horizontal_gridlines",
                    "Show Row Gridlines": "horizontal_gridlines",  # Alias for backward compatibility
                    "Row Dividers": "horizontal_gridlines",
                    "Vertical Gridline Years": "vertical_gridline_years",
                    "Vertical Gridline Months": "vertical_gridline_months",
                    "Vertical Gridline Weeks": "vertical_gridline_weeks",
                    "Vertical Gridline Days": "vertical_gridline_days",
                    "Row Numbers": "show_row_numbers",
                    "Show Row Numbers": "show_row_numbers",  # Alias for backward compatibility
                    "Chart Start Date": "chart_start_date",
                    "Chart End Date": "chart_end_date"
                }
                
                field_name = field_map.get(key)
                if field_name:
                    if field_name == "margins":
                        # Handle margins tuple
                        if "margins" not in data:
                            data["margins"] = [0, 0, 0, 0]
                        margin_map = {
                            "Margin Top": 0,
                            "Margin Right": 1,
                            "Margin Bottom": 2,
                            "Margin Left": 3
                        }
                        margin_idx = margin_map.get(key)
                        if margin_idx is not None:
                            data["margins"][margin_idx] = value
                    else:
                        data[field_name] = value
        
        # Convert margins list to tuple
        if "margins" in data and isinstance(data["margins"], list):
            data["margins"] = tuple(data["margins"])
        
        return data
    
    def _read_tasks_sheet(self, ws) -> List[Task]:
        """Read Tasks worksheet and return list of Task objects.
        
        Auto-assigns sequential IDs and row numbers to tasks that don't have them.
        This allows users to paste simple task lists (Name, Start Date, Finish Date)
        without needing to manually specify IDs and row numbers.
        """
        tasks = []
        headers = None
        task_data_list = []  # Collect all task data first
        
        # First pass: collect all task data from Excel
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # Header row
                headers = [str(cell).strip() if cell else "" for cell in row]
                continue
            
            if not any(row):  # Skip empty rows
                continue
            
            # Create task dict from row data
            task_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    header = headers[col_idx]
                    
                    # Map Excel headers to task fields
                    if header == "ID":
                        task_data["task_id"] = int(value) if value is not None else 0
                    elif header == "Row":
                        # Only set row_number if value is provided, otherwise leave unset for auto-assignment
                        if value is not None:
                            try:
                                task_data["row_number"] = int(value)
                            except (ValueError, TypeError):
                                # Invalid value, leave unset for auto-assignment
                                pass
                    elif header == "Name":
                        task_data["task_name"] = str(value) if value is not None else ""
                    elif header == "Start Date":
                        # Convert from display format to internal format
                        # Handle both string dates and Excel date serial numbers
                        try:
                            if value:
                                if isinstance(value, datetime):
                                    # Excel date serial number converted to datetime
                                    task_data["start_date"] = value.strftime("%Y-%m-%d")
                                else:
                                    # String date - convert from display format
                                    task_data["start_date"] = display_to_internal_date(str(value))
                            else:
                                task_data["start_date"] = ""
                        except (ValueError, AttributeError):
                            task_data["start_date"] = ""
                    elif header == "Finish Date":
                        # Convert from display format to internal format
                        # Handle both string dates and Excel date serial numbers
                        try:
                            if value:
                                if isinstance(value, datetime):
                                    # Excel date serial number converted to datetime
                                    task_data["finish_date"] = value.strftime("%Y-%m-%d")
                                else:
                                    # String date - convert from display format
                                    task_data["finish_date"] = display_to_internal_date(str(value))
                            else:
                                task_data["finish_date"] = ""
                        except (ValueError, AttributeError):
                            task_data["finish_date"] = ""
                    elif header == "Label" or header == "Label Display" or header == "Label Content":
                        # Support both old "Label Display" (Yes/No) and new "Label Content" (None/Name only/Date only/Name and Date)
                        value_str = str(value).strip() if value is not None else ""
                        if value_str in ["None", "Name only", "Date only", "Name and Date"]:
                            # New format
                            task_data["label_content"] = value_str
                            task_data["label_hide"] = "No" if value_str == "None" else "Yes"  # Keep for backward compatibility
                        elif value_str.lower() in ["yes", "no"]:
                            # Old format - migrate to new format
                            task_data["label_hide"] = value_str
                            task_data["label_content"] = "None" if value_str.lower() == "no" else "Name only"
                        else:
                            # Default
                            task_data["label_content"] = "Name only"
                            task_data["label_hide"] = "Yes"
                    elif header == "Placement" or header == "Label Placement":
                        task_data["label_placement"] = str(value) if value is not None else "Outside"
                    elif header == "Offset" or header == "Label Offset" or header == "Label Horizontal Offset":
                        task_data["label_horizontal_offset"] = float(value) if value is not None else 0.0
                    elif header == "Fill Color":
                        task_data["fill_color"] = str(value) if value is not None else "blue"
                    elif header == "Is Milestone":
                        task_data["is_milestone"] = str(value).strip().lower() in ["yes", "true", "1"]
                    elif header == "Label Alignment":
                        task_data["label_alignment"] = str(value) if value is not None else "Centre"
                    elif header == "Label Text Colour":
                        task_data["label_text_colour"] = str(value) if value is not None else "black"
            
            # Check if row has valid data (at least Name or dates)
            has_name = task_data.get("task_name", "").strip()
            has_start_date = task_data.get("start_date", "").strip()
            has_finish_date = task_data.get("finish_date", "").strip()
            
            if has_name or has_start_date or has_finish_date:
                # Row has valid data, add to list for processing
                task_data_list.append(task_data)
        
        # Second pass: auto-assign IDs and row numbers, then create tasks
        # Find max existing ID to start auto-assignment from
        max_id = 0
        for task_data in task_data_list:
            task_id = task_data.get("task_id", 0)
            if task_id > max_id:
                max_id = task_id
        
        # Auto-assign IDs and row numbers
        next_id = max_id + 1
        next_row = 1
        
        for task_data in task_data_list:
            # Auto-assign ID if missing or invalid
            if "task_id" not in task_data or task_data.get("task_id", 0) <= 0:
                task_data["task_id"] = next_id
                next_id += 1
            
            # Auto-assign row number if missing or invalid
            if "row_number" not in task_data or task_data.get("row_number", 0) <= 0:
                task_data["row_number"] = next_row
                next_row += 1
            
            # Create task if we have a valid ID
            if task_data.get("task_id", 0) > 0:
                try:
                    task = Task.from_dict(task_data)
                    tasks.append(task)
                except (KeyError, ValueError) as e:
                    # Skip invalid tasks
                    logging.warning(f"Skipping invalid task row: {e}")
                    continue
        
        return tasks
    
    def _read_links_sheet(self, ws) -> List[Link]:
        """Read Links worksheet and return list of Link objects using header-based mapping.
        
        This method uses header-based mapping instead of positional arrays to avoid
        index brittleness when columns are added, removed, or reordered.
        """
        links = []
        headers = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # Header row
                headers = [str(cell).strip() if cell else "" for cell in row]
                continue
            
            if not any(row):  # Skip empty rows
                continue
            
            # Create link dict from row data using header mapping
            link_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    header = headers[col_idx]
                    
                    # Map Excel headers to link fields
                    if header == "ID":
                        link_data["link_id"] = int(value) if value is not None else 0
                    elif header == "From Task ID":
                        link_data["from_task_id"] = int(value) if value is not None else 0
                    elif header == "From Task Name":
                        link_data["from_task_name"] = str(value) if value is not None else ""
                    elif header == "To Task ID":
                        link_data["to_task_id"] = int(value) if value is not None else 0
                    elif header == "To Task Name":
                        link_data["to_task_name"] = str(value) if value is not None else ""
                    elif header == "Line Color":
                        link_data["line_color"] = str(value) if value is not None else "black"
                    elif header == "Line Style":
                        link_data["line_style"] = str(value) if value is not None else "solid"
                    elif header == "Link Routing":
                        link_data["link_routing"] = str(value) if value is not None else "auto"
            
            # Only create link if we have required fields
            if ("link_id" in link_data and link_data["link_id"] > 0 and
                "from_task_id" in link_data and link_data["from_task_id"] > 0 and
                "to_task_id" in link_data and link_data["to_task_id"] > 0):
                try:
                    link = Link.from_dict(link_data)
                    links.append(link)
                except (KeyError, ValueError) as e:
                    # Skip invalid links
                    continue
        
        return links
    
    def _read_text_boxes_sheet(self, ws) -> List[TextBox]:
        """Read Text Boxes worksheet and return list of TextBox objects."""
        text_boxes = []
        headers = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # Header row
                headers = [str(cell).strip() if cell else "" for cell in row]
                continue
            
            if not any(row):  # Skip empty rows
                continue
            
            # Create textbox dict from row data using header mapping
            textbox_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    header = headers[col_idx]
                    
                    # Map Excel headers to textbox fields
                    if header == "ID":
                        textbox_data["textbox_id"] = int(value) if value is not None else 0
                    elif header == "X":
                        textbox_data["x"] = int(value) if value is not None else 0
                    elif header == "Y":
                        textbox_data["y"] = int(value) if value is not None else 0
                    elif header == "Width":
                        textbox_data["width"] = int(value) if value is not None else 100
                    elif header == "Height":
                        textbox_data["height"] = int(value) if value is not None else 50
                    elif header == "Text Align":
                        textbox_data["text_align"] = str(value) if value is not None else "Center"
                    elif header == "Vertical Align":
                        textbox_data["vertical_align"] = str(value) if value is not None else "Middle"
                    elif header == "Text":
                        textbox_data["text"] = str(value) if value is not None else ""
            
            # Only create textbox if we have required fields
            if ("textbox_id" in textbox_data and textbox_data["textbox_id"] > 0 and
                "x" in textbox_data and "y" in textbox_data and
                "width" in textbox_data and "height" in textbox_data):
                try:
                    textbox = TextBox.from_dict(textbox_data)
                    text_boxes.append(textbox)
                except (KeyError, ValueError) as e:
                    # Skip invalid textboxes
                    continue
        
        return text_boxes
    
    def _read_typography_sheet(self, ws) -> Dict[str, Any]:
        """Read Typography worksheet and return dict of chart_config fields."""
        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row[0] and row[1] is not None:
                key = str(row[0]).strip()
                value = row[1]
                
                # Map Excel field names to chart_config field names
                field_map = {
                    "Font Family": "font_family",
                    "Task Font Size": "task_font_size",
                    "Scale Font Size": "scale_font_size",
                    "Header & Footer Font Size": "header_footer_font_size",
                    "Row Number Font Size": "row_number_font_size",
                    "Text Box Font Size": "text_box_font_size",
                    "Swimlane Font Size": "swimlane_font_size",
                    "Scale Vertical Alignment Factor": "scale_vertical_alignment_factor",
                    "Task Vertical Alignment Factor": "task_vertical_alignment_factor",
                    "Row Number Vertical Alignment Factor": "row_number_vertical_alignment_factor",
                    "Header & Footer Vertical Alignment Factor": "header_footer_vertical_alignment_factor",
                    "Swimlane Vertical Alignment Factor": "swimlane_vertical_alignment_factor",  # Old format for backward compatibility
                    "Swimlane Top Vertical Alignment Factor": "swimlane_top_vertical_alignment_factor",
                    "Swimlane Bottom Vertical Alignment Factor": "swimlane_bottom_vertical_alignment_factor"
                }
                
                field_name = field_map.get(key)
                if field_name:
                    # Convert value based on field type
                    if field_name == "font_family":
                        data[field_name] = str(value) if value is not None else "Arial"
                    elif "font_size" in field_name:
                        try:
                            data[field_name] = int(value) if value is not None else 10
                        except (ValueError, TypeError):
                            data[field_name] = 10
                    elif "alignment_factor" in field_name:
                        try:
                            data[field_name] = float(value) if value is not None else 0.7
                        except (ValueError, TypeError):
                            data[field_name] = 0.7
        
        # Handle backward compatibility: if old swimlane_vertical_alignment_factor exists, use it for both
        if "swimlane_vertical_alignment_factor" in data and "swimlane_top_vertical_alignment_factor" not in data:
            old_factor = data["swimlane_vertical_alignment_factor"]
            data["swimlane_top_vertical_alignment_factor"] = old_factor
            data["swimlane_bottom_vertical_alignment_factor"] = old_factor
        
        return data
    
    def _read_table_sheet(self, ws) -> List[List[str]]:
        """Read a table worksheet (legacy format for backward compatibility)."""
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(row):  # Only add non-empty rows
                data.append([str(cell) if cell is not None else "" for cell in row])
        return data[1:] if len(data) > 1 else []  # Skip header row

